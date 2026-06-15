"""
P2P_Registry_Sync.py

Syncs Yardi P2P milestone data into the selected registry workbook.

Columns filled / updated:
  P2P Status  <- ir/milestones or po/milestones  (only if column exists)
  P.O. #      <- IRGrid API                      (only fills empty cells)
  P2P IR #    <- IRGrid API                      (only fills empty cells)
  Ch.#        <- ir/milestones Paid milestone     (only fills empty cells)
  Ch. Date    <- ir/milestones Paid milestone     (only fills empty cells)

All column positions are detected dynamically from the header row so the
script keeps working even if columns are reordered later.

What it NEVER does:
  - Delete or move any row (empty rows kept for manager's visual layout)
  - Insert the P2P Status column if it is missing
  - Overwrite PO, IR, Ch.#, Ch. Date cells that already have a value
  - Change formatting, column widths, or styles on existing cells
  - Save changes without first backing up the selected input workbook

Run:  python P2P_Registry_Sync.py
"""

import json
import sys
import os
import shutil
import time
import base64
from datetime import datetime

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright

# ---------------------------------------------------------------------------
# Config
# ---------------------------------------------------------------------------

def get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))

BASE_DIR      = get_base_dir()
CONFIG_PATH   = os.path.join(BASE_DIR, "p2p_private_config.json")
SESSION_DIR   = os.path.join(BASE_DIR, "browser_session")
DEFAULT_REGISTRY_PATH = os.path.join(BASE_DIR, "01. Cheque Registry_Year 2026.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Output Excel File")


def load_private_config():
    if not os.path.isfile(CONFIG_PATH):
        raise FileNotFoundError(
            "Missing private config file: p2p_private_config.json. "
            "Copy p2p_private_config.example.json, fill in your private values, "
            "and keep p2p_private_config.json out of public repos."
        )
    with open(CONFIG_PATH, "r", encoding="utf-8") as config_file:
        return json.load(config_file)


PRIVATE_CONFIG = load_private_config()

if not os.path.isfile(DEFAULT_REGISTRY_PATH):
    for filename in os.listdir(BASE_DIR):
        if filename.startswith("~$"):
            continue
        if filename.lower().endswith((".xlsx", ".xlsm")):
            DEFAULT_REGISTRY_PATH = os.path.join(BASE_DIR, filename)
            break

LOGIN_URL        = PRIVATE_CONFIG["login_url"]
API_BASE         = PRIVATE_CONFIG["api_base"].rstrip("/")
P2P_DATABASE     = PRIVATE_CONFIG["database"]
P2P_ORIGIN       = PRIVATE_CONFIG["origin"]
P2P_REFERER      = PRIVATE_CONFIG["referer"]
IRGRID_URL       = f"{API_BASE}/IRTileData/IRGrid"
IR_MILESTONE_URL = f"{API_BASE}/ir/milestones"
PO_MILESTONE_URL = f"{API_BASE}/po/milestones"
POGRID_URL       = f"{API_BASE}/po/charts/data"
BULK_IRGRID_RECORD_LIMIT = 1000
BULK_PO_RECORD_LIMIT = 1000
STATUS_MODE = "detailed"  # detailed, hybrid, or fast
VALID_STATUS_MODES = ("detailed", "hybrid", "fast")

# Header names in the registry — used to find columns dynamically
H_CH_DATE   = "Ch. Date (DD/MM/YYYY)"
H_CH_NUM    = "Ch.#"
H_STATUS    = "P2P Status"
H_PO        = "P.O. #"
H_IR        = "P2P IR #"
H_INVOICE   = "Invoice No"

IRGRID_INVOICE_FIELDS = ("Invoice", "InvoiceNumber", "InvoiceNo", "InvoiceNum", "InvNumber")
IRGRID_IR_FIELDS = ("Inv_hmy", "IR_hmy", "IR", "InvoiceNumber")
IRGRID_PO_FIELDS = ("PO_hmy", "PO", "PONumber", "POId")
IRGRID_STATUS_FIELDS = ("Status", "workflowStepName", "WorkflowStatus")

POGRID_PO_FIELDS = ("PO", "POCode", "PONumber", "PONum", "PONo")
POGRID_STATUS_FIELDS = ("WorkflowStatus", "Status", "POStatus")
MILESTONE_DATE_FIELDS = (
    "RevisedDate", "DtCompleted", "Date", "CompletedDate", "DateCompleted",
    "DtComplete", "DtCompletion", "CompletionDate", "ActualDate",
)


def clean_dropped_path(raw_path):
    """Normalize a path pasted by terminal drag-and-drop."""
    path = str(raw_path or "").strip()
    if path.startswith("& "):
        path = path[2:].strip()
    return path.strip().strip('"').strip("'")


def get_registry_path():
    """
    Accept an Excel path from argv or terminal drag-and-drop.
    Pressing Enter keeps the original default workbook path.
    """
    if len(sys.argv) > 1:
        return clean_dropped_path(" ".join(sys.argv[1:]))

    print("Drag and drop the Excel registry file into this terminal, then press Enter.")
    print(f"Press Enter without a file to use: {os.path.basename(DEFAULT_REGISTRY_PATH)}")
    selected = clean_dropped_path(input("Excel file: "))
    return selected or DEFAULT_REGISTRY_PATH


def get_status_mode():
    """
    Let the user override STATUS_MODE at runtime.
    Pressing Enter keeps the default detailed mode.
    """
    print("\nP2P status mode:")
    print("  detailed = exact workflow milestone status (default)")
    print("  hybrid   = fewer IR milestone calls, detailed PO milestones")
    print("  fast     = bulk status only, fastest but less detailed")

    selected = input(f"Mode [{STATUS_MODE}]: ").strip().lower()
    if not selected:
        return STATUS_MODE
    if selected in VALID_STATUS_MODES:
        return selected

    print(f"Unknown mode '{selected}' - using default: {STATUS_MODE}")
    return STATUS_MODE


def get_backup_path(registry_path):
    """Create a timestamped backup path for the original input workbook."""
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _, ext = os.path.splitext(os.path.basename(registry_path))
    return os.path.join(OUTPUT_DIR, f"back up_{ts}{ext}")


def backup_existing_workbook(registry_path):
    backup_path = get_backup_path(registry_path)
    shutil.copy2(registry_path, backup_path)
    return backup_path

# ---------------------------------------------------------------------------
# Auth helpers
# ---------------------------------------------------------------------------

def get_auth_headers(context):
    all_cookies = {c["name"]: c["value"] for c in context.cookies()}
    jwt_token = all_cookies.get(".JWTAUTH")
    if not jwt_token:
        return None
    try:
        payload_b64 = jwt_token.split(".")[1]
        payload_b64 += "=" * (4 - len(payload_b64) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        if payload.get("exp", 0) < time.time():
            return None
        login_guid = payload.get("UserLoginGUID", "")
        role = payload.get(
            "http://schemas.microsoft.com/ws/2008/06/identity/claims/role", ""
        )
    except Exception:
        return None
    return {
        "Authorization":   f"Bearer {jwt_token}",
        "Database":        P2P_DATABASE,
        "Loginguid":       login_guid,
        "Role":            role,
        "Content-Type":    "application/json;charset=UTF-8",
        "Accept":          "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Origin":          P2P_ORIGIN,
        "Referer":         P2P_REFERER,
        "Sec-Fetch-Dest":  "empty",
        "Sec-Fetch-Mode":  "cors",
        "Sec-Fetch-Site":  "same-origin",
    }

def wait_for_jwt(context, timeout_seconds=300):
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        headers = get_auth_headers(context)
        if headers:
            return headers
        time.sleep(1)
    return None

def ensure_authenticated(page, context):
    headers = get_auth_headers(context)
    if headers:
        print("Valid session found — skipping login.\n")
        return headers
    print("\n  Browser opened — please log in to continue.")
    print("  After logging in, click 'Elevate (Production)' to open P2P.")
    print("  The script will continue automatically once you are in.\n")
    page.goto(LOGIN_URL)
    headers = wait_for_jwt(context, timeout_seconds=300)
    if headers:
        print("  Login detected — continuing.\n")
        return headers
    raise TimeoutError("Login timed out after 5 minutes.")

# ---------------------------------------------------------------------------
# API calls
# ---------------------------------------------------------------------------

def lookup_ir_po(invoice, page, headers):
    """Returns (ir_id, po_id) as ints. Either can be None (contract = no PO)."""
    resp = page.request.post(
        IRGRID_URL,
        params={
            "NumRecords":      2,
            "search":          str(invoice),
            "enddate":         "",
            "invoicedatefrom": "",
            "invoicedateto":   "",
            "startdate":       "",
        },
        data=json.dumps([]),
        headers=headers,
    )
    records = resp.json().get("Records", [])
    if not records:
        return None, None

    invoice_key = normalize_invoice_key(invoice)
    matching_record = None
    for rec in records:
        for field in IRGRID_INVOICE_FIELDS:
            if normalize_invoice_key(rec.get(field)) == invoice_key:
                matching_record = rec
                break
        if matching_record:
            break

    if not matching_record:
        return None, None

    return (
        _first_int_field(matching_record, IRGRID_IR_FIELDS),
        _first_int_field(matching_record, IRGRID_PO_FIELDS),
    )


def normalize_invoice_key(value):
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0"):
        text = text[:-2]
    return text.lower()


def _first_int_field(record, fields):
    for field in fields:
        if field in record:
            value = to_int_id(record.get(field))
            if value is not None:
                return value
    return None


def _first_text_field(record, fields):
    for field in fields:
        if field in record and not is_empty(record.get(field)):
            return str(record.get(field)).strip()
    return ""


def _unwrap_irgrid_records(raw):
    if isinstance(raw, dict):
        records = raw.get("Records") or raw.get("records") or raw.get("Data") or raw.get("data")
        return records if isinstance(records, list) else []
    if isinstance(raw, list):
        return raw
    return []


def preload_irgrid(page, headers):
    """
    Pull the IRGrid once, matching the website's broad "Any" search.
    Returns a local invoice index so per-row IRGrid searches are avoided.
    """
    print("\nPreloading IRGrid records for local lookup...")
    resp = page.request.post(
        IRGRID_URL,
        params={
            "NumRecords":      BULK_IRGRID_RECORD_LIMIT,
            "search":          "",
            "enddate":         "",
            "invoicedatefrom": "",
            "invoicedateto":   "",
            "startdate":       "",
        },
        data=json.dumps([]),
        headers=headers,
    )
    records = _unwrap_irgrid_records(resp.json())
    index = {}

    for record in records:
        for field in IRGRID_INVOICE_FIELDS:
            if field not in record or is_empty(record.get(field)):
                continue
            key = normalize_invoice_key(record.get(field))
            if key:
                index.setdefault(key, []).append((record, field))

    print(f"  IRGrid records loaded : {len(records)}")
    print(f"  Local lookup keys     : {len(index)}")
    return index


def lookup_ir_po_from_bulk(invoice, irgrid_index):
    """Returns (ir_id, po_id, matched_field, status), or empty values."""
    matches = irgrid_index.get(normalize_invoice_key(invoice), [])
    if not matches:
        return None, None, None, ""

    # Prefer vendor invoice matches over internal IR-number matches.
    field_priority = {field: idx for idx, field in enumerate(IRGRID_INVOICE_FIELDS)}
    record, matched_field = min(
        matches,
        key=lambda item: field_priority.get(item[1], len(field_priority)),
    )
    return (
        _first_int_field(record, IRGRID_IR_FIELDS),
        _first_int_field(record, IRGRID_PO_FIELDS),
        matched_field,
        _first_text_field(record, IRGRID_STATUS_FIELDS),
    )


def _unwrap_grid_records(raw):
    if isinstance(raw, dict):
        records = raw.get("Records") or raw.get("records") or raw.get("Data") or raw.get("data")
        return records if isinstance(records, list) else []
    if isinstance(raw, list):
        return raw
    return []


def preload_pogrid(page, headers):
    """Pull the PO grid once so PO-only rows can use local status data."""
    print("\nPreloading PO records for local lookup...")
    resp = page.request.post(
        POGRID_URL,
        params={
            "NumRecords":      BULK_PO_RECORD_LIMIT,
            "dateFrom":        "",
            "dateTo":          "",
            "showMarketplace": 0,
        },
        data=json.dumps([]),
        headers=headers,
    )
    records = _unwrap_grid_records(resp.json())
    index = {}

    for record in records:
        for field in POGRID_PO_FIELDS:
            if field not in record or is_empty(record.get(field)):
                continue
            key = normalize_invoice_key(record.get(field))
            if key:
                index.setdefault(key, []).append((record, field))

    print(f"  PO records loaded     : {len(records)}")
    print(f"  Local PO lookup keys  : {len(index)}")
    return index


def lookup_po_status_from_bulk(po_id, pogrid_index):
    """Returns broad PO status from the PO grid, or ''. """
    matches = pogrid_index.get(normalize_invoice_key(po_id), [])
    if not matches:
        return ""
    record, _ = matches[0]
    status = _first_text_field(record, POGRID_STATUS_FIELDS)
    return f"PO: {status}" if status else ""


def format_ir_bulk_status(status):
    status = str(status or "").strip()
    if not status:
        return ""
    return status if status.startswith("IR:") else f"IR: {status}"


def should_use_detailed_ir_status(status_mode, need_ch, need_chd):
    if status_mode == "detailed":
        return True
    if status_mode == "hybrid":
        return need_ch or need_chd
    return False


def should_use_detailed_po_status(status_mode):
    return status_mode in ("detailed", "hybrid")


def _unwrap_milestones(raw):
    """Handle both bare list and wrapped {"data": [...]} response."""
    if isinstance(raw, list):
        return raw
    if isinstance(raw, dict):
        return (
            raw.get("data") or raw.get("Records") or raw.get("Data")
            or next((v for v in raw.values() if isinstance(v, list)), [])
        )
    return []


def _format_milestone_date(date_str):
    """Convert 'YYYY-MM-DDTHH:MM:SS' to 'D/M/YYYY'. Returns '' on failure."""
    if not date_str or date_str in ("None", ""):
        return ""
    try:
        dt = datetime.strptime(date_str.split("T")[0], "%Y-%m-%d")
        return f"{dt.day}/{dt.month}/{dt.year}"
    except ValueError:
        return ""


def _parse_milestone_date(milestone, preferred_date_field):
    fields = (preferred_date_field,) + tuple(
        field for field in MILESTONE_DATE_FIELDS if field != preferred_date_field
    )
    for field in fields:
        raw_date = milestone.get(field, "")
        if not raw_date or str(raw_date) in ("None", ""):
            continue
        try:
            return datetime.strptime(str(raw_date).split("T")[0], "%Y-%m-%d")
        except ValueError:
            continue
    return datetime.min


def _is_truthy_api_flag(value):
    return value is True or str(value).strip().lower() in ("1", "true", "yes")


def _has_parseable_date(milestone, field):
    raw_date = milestone.get(field, "")
    if not raw_date or str(raw_date) in ("None", ""):
        return False
    try:
        datetime.strptime(str(raw_date).split("T")[0], "%Y-%m-%d")
        return True
    except ValueError:
        return False


def _is_completed(milestone, date_field=None):
    if _is_truthy_api_flag(milestone.get("IsComplete")):
        return True
    if date_field and _has_parseable_date(milestone, date_field):
        return True
    return _parse_milestone_date(milestone, date_field or "Date") != datetime.min


def _looks_current(milestone):
    current_keys = (
        "IsInProgress", "InProgress", "IsCurrent", "Current", "IsCurrentStep",
        "CurrentStep", "IsActive", "Active", "IsPending", "Pending",
    )
    if any(_is_truthy_api_flag(milestone.get(key)) for key in current_keys):
        return True

    for key, value in milestone.items():
        key_text = str(key).lower()
        value_text = str(value).strip().lower()
        if ("current" in key_text or "progress" in key_text) and _is_truthy_api_flag(value):
            return True
        if key_text in ("status", "state", "class", "cssclass") and (
            "current" in value_text or "progress" in value_text or "pending" in value_text
        ):
            return True

    return False


def _seq_value(milestone):
    try:
        return int(float(milestone.get("Seq", 0) or 0))
    except (ValueError, TypeError):
        return 0


def _milestone_sort_key(milestone, preferred_date_field, index=0):
    return (_parse_milestone_date(milestone, preferred_date_field), _seq_value(milestone), index)


def _latest_milestone(milestones, preferred_date_field):
    return max(
        enumerate(milestones),
        key=lambda item: _milestone_sort_key(item[1], preferred_date_field, item[0]),
    )[1]


def _next_milestone(milestones, preferred_date_field):
    return min(
        enumerate(milestones),
        key=lambda item: _milestone_sort_key(item[1], preferred_date_field, item[0]),
    )[1]


def _milestone_date(milestone, preferred_date_field):
    fields = (preferred_date_field,) + tuple(
        field for field in MILESTONE_DATE_FIELDS if field != preferred_date_field
    )
    for field in fields:
        formatted = _format_milestone_date(str(milestone.get(field, "")))
        if formatted:
            return formatted
    return ""


def _milestone_title(milestone):
    for field in ("Title", "Name", "StepName", "MilestoneName", "Description"):
        title = str(milestone.get(field, "")).strip()
        if title:
            return title
    return "Submitted"


def _current_milestone(milestones, date_field):
    """
    Return (title, formatted_date) for the current visible milestone.
    In-progress milestones are preferred because they reflect where the item
    currently sits in the workflow. If none are in progress, use the latest
    completed milestone.
    date_field is 'RevisedDate' for IR milestones, 'DtCompleted' for PO milestones.
    """
    current = [m for m in milestones if _looks_current(m)]
    if current:
        m = _latest_milestone(current, date_field)
        return _milestone_title(m), _milestone_date(m, date_field)

    completed = [m for m in milestones if _is_completed(m, date_field)]
    if completed:
        latest = _latest_milestone(completed, date_field)
        title  = _milestone_title(latest)
        date   = _milestone_date(latest, date_field)
        return title, date

    next_steps = [m for m in milestones if not _is_completed(m, date_field)]
    if not next_steps:
        return "Submitted", ""
    m = _next_milestone(next_steps, date_field)
    return _milestone_title(m), _milestone_date(m, date_field)


def _format_status(title, date, prefix=""):
    if title.strip().lower() == "paid" and date:
        status = f"Cheque has issued on {date}"
        return f"{prefix}{status}" if prefix else status
    status = f"{title}  {date}".strip() if date else title
    return f"{prefix}{status}" if prefix else status


def lookup_ir_milestone(ir_id, page, headers):
    """
    Returns:
      status_str  : current visible P2P milestone status
      check_num   : str or None  (only when Paid milestone exists)
      check_date  : str D/M/YYYY or None
    """
    resp = page.request.post(
        IR_MILESTONE_URL,
        params={"Id": ir_id},
        data=json.dumps([]),
        headers=headers,
    )
    milestones = _unwrap_milestones(resp.json())
    if not milestones:
        return "No milestones", None, None

    title, date = _current_milestone(milestones, "RevisedDate")
    status_str  = _format_status(title, date, prefix="IR: ")

    # Also extract cheque data from the Paid milestone if present
    paid = next(
        (m for m in milestones
         if str(m.get("Title", "")).strip().lower() == "paid"
         and _is_completed(m, "RevisedDate")),
        None,
    )
    check_num  = None
    check_date = None
    if paid:
        check_num  = str(paid.get("CheckNum", "")).strip() or None
        check_date = _format_milestone_date(str(paid.get("RevisedDate", ""))) or None

    return status_str, check_num, check_date


def lookup_po_milestone(po_id, page, headers):
    """Returns status_str for a PO-only row (no IR yet)."""
    resp = page.request.post(
        PO_MILESTONE_URL,
        params={"Id": po_id},
        data=json.dumps([]),
        headers=headers,
    )
    milestones = _unwrap_milestones(resp.json())
    if not milestones:
        return "PO: No milestones"
    title, date = _current_milestone(milestones, "DtCompleted")
    status_str  = _format_status(title, date, prefix="PO: ")
    return status_str

# ---------------------------------------------------------------------------
# Cell helpers
# ---------------------------------------------------------------------------

def is_empty(val):
    if val is None:
        return True
    return str(val).strip().lower() in ("", "nan", "none")


def to_int_id(val):
    if is_empty(val):
        return None
    try:
        return int(float(str(val).strip()))
    except (ValueError, TypeError):
        return None


def is_cheque_issued(ch_num, ch_date, status):
    """Rows with issued cheque data are final and do not need P2P refresh."""
    if not is_empty(ch_num) and not is_empty(ch_date):
        return True
    status_text = str(status or "").strip().lower()
    return "cheque has issued" in status_text or status_text.startswith("ir: paid")


def get_col_map(ws):
    """Scan header row (row 1) and return {header_name: col_index}."""
    return {
        str(cell.value).strip(): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def sync_sheet(ws, sheet_name, page, auth_headers, status_mode, irgrid_index, pogrid_index,
               ir_cache, cheque_cache, po_status_cache):
    """
    Process one worksheet. Fills P2P Status only when the column exists,
    and fills empty PO, IR, Ch.#, Ch. Date cells.
    Returns stats dict.
    """
    col_map = get_col_map(ws)

    col_status  = col_map.get(H_STATUS)
    col_ch_date = col_map.get(H_CH_DATE)
    col_ch_num  = col_map.get(H_CH_NUM)
    col_po      = col_map.get(H_PO)
    col_ir      = col_map.get(H_IR)
    col_invoice = col_map.get(H_INVOICE)

    if not col_status:
        print(f"  '{H_STATUS}' column not found - status updates skipped.")

    if not col_invoice:
        print(f"  '{H_INVOICE}' column not found — skipping sheet.")
        return {"filled_po": 0, "filled_ir": 0, "filled_cheque": 0,
                "status_updated": 0, "not_found": 0, "bulk_hits": 0,
                "fallback_lookups": 0, "cheque_issued_skipped": 0}

    stats = {"filled_po": 0, "filled_ir": 0, "filled_cheque": 0,
             "status_updated": 0, "not_found": 0, "bulk_hits": 0,
             "fallback_lookups": 0, "cheque_issued_skipped": 0}

    for row_idx in range(2, ws.max_row + 1):
        invoice_val = ws.cell(row=row_idx, column=col_invoice).value
        if is_empty(invoice_val):
            continue   # blank / spacing row — skip silently

        invoice = str(invoice_val).strip()
        label   = f"  [{sheet_name}] Row {row_idx:>4} | {invoice:<18}"

        # Read current cell values
        po_val  = ws.cell(row=row_idx, column=col_po).value  if col_po  else None
        ir_val  = ws.cell(row=row_idx, column=col_ir).value  if col_ir  else None
        ch_val  = ws.cell(row=row_idx, column=col_ch_num).value if col_ch_num else None
        chd_val = ws.cell(row=row_idx, column=col_ch_date).value if col_ch_date else None
        status_val = ws.cell(row=row_idx, column=col_status).value if col_status else None

        if is_cheque_issued(ch_val, chd_val, status_val):
            stats["cheque_issued_skipped"] += 1
            continue

        need_po  = is_empty(po_val)
        need_ir  = is_empty(ir_val)
        need_ch  = is_empty(ch_val)
        need_chd = is_empty(chd_val)

        current_ir_id = to_int_id(ir_val)
        current_po_id = to_int_id(po_val)

        # ── Step 1: IRGrid lookup if PO or IR is missing ────────────
        if need_po or need_ir:
            cache_key = normalize_invoice_key(invoice)
            if cache_key not in ir_cache:
                ir_id, po_id, matched_field, bulk_status = lookup_ir_po_from_bulk(invoice, irgrid_index)
                if matched_field:
                    ir_cache[cache_key] = (ir_id, po_id, f"bulk:{matched_field}", bulk_status)
                else:
                    print(f"{label}  IRGrid fallback ...", end="  ", flush=True)
                    stats["fallback_lookups"] += 1
                    try:
                        ir_id, po_id = lookup_ir_po(invoice, page, auth_headers)
                        ir_cache[cache_key] = (ir_id, po_id, "fallback", "")
                        time.sleep(0.3)
                    except Exception as e:
                        print(f"error: {e}")
                        ir_cache[cache_key] = (None, None, "error", "")

            ir_id, po_id, lookup_source, bulk_status = ir_cache[cache_key]
            if lookup_source.startswith("bulk:"):
                stats["bulk_hits"] += 1

            if ir_id is None and po_id is None:
                if lookup_source == "fallback":
                    print("not found in P2P")
                else:
                    print(f"{label}  not found in P2P")
                stats["not_found"] += 1
                # Still try PO milestone if PO is manually entered
                if current_po_id:
                    pass   # handled in status step below
                else:
                    if col_status:
                        ws.cell(row=row_idx, column=col_status).value = "Not in P2P"
                        stats["status_updated"] += 1
                    continue
            else:
                parts = []
                if need_po:
                    if po_id is not None:
                        ws.cell(row=row_idx, column=col_po).value = po_id
                        stats["filled_po"] += 1
                        current_po_id = po_id
                        parts.append(f"PO={po_id}")
                    else:
                        parts.append("PO=N/A(contract)")
                if need_ir and ir_id is not None:
                    ws.cell(row=row_idx, column=col_ir).value = ir_id
                    stats["filled_ir"] += 1
                    current_ir_id = ir_id
                    parts.append(f"IR={ir_id}")
                if parts:
                    print(f"{label}  {', '.join(parts)}  ({lookup_source})")

        # ── Step 2: Milestone lookup — IR path (preferred) ──────────
        need_ir_milestone = (
            should_use_detailed_ir_status(status_mode, need_ch, need_chd)
            or (need_ch and col_ch_num)
            or (need_chd and col_ch_date)
        )

        if current_ir_id and need_ir_milestone:
            if current_ir_id not in cheque_cache:
                print(f"{label}  IR milestones ...", end="  ", flush=True)
                try:
                    status_str, chk_num, chk_date = lookup_ir_milestone(
                        current_ir_id, page, auth_headers
                    )
                    cheque_cache[current_ir_id] = (status_str, chk_num, chk_date)
                    time.sleep(0.3)
                except Exception as e:
                    print(f"error: {e}")
                    cheque_cache[current_ir_id] = ("Error", None, None)

            status_str, chk_num, chk_date = cheque_cache[current_ir_id]

            if col_status:
                ws.cell(row=row_idx, column=col_status).value = status_str
                stats["status_updated"] += 1

            # Fill cheque fields only if empty
            if need_ch and chk_num and col_ch_num:
                try:
                    ws.cell(row=row_idx, column=col_ch_num).value = int(chk_num)
                except ValueError:
                    ws.cell(row=row_idx, column=col_ch_num).value = chk_num
                stats["filled_cheque"] += 1

            if need_chd and chk_date and col_ch_date:
                ws.cell(row=row_idx, column=col_ch_date).value = chk_date

            print(f"{label}  {status_str}")

        # ── Step 3: PO milestone path (no IR yet) ───────────────────
        elif current_ir_id and col_status:
            status_str = format_ir_bulk_status(locals().get("bulk_status", ""))
            if status_str:
                ws.cell(row=row_idx, column=col_status).value = status_str
                stats["status_updated"] += 1
                print(f"{label}  {status_str}  (bulk status)")

        elif not current_ir_id and current_po_id and should_use_detailed_po_status(status_mode):
            if current_po_id not in po_status_cache:
                print(f"{label}  PO milestones ...", end="  ", flush=True)
                try:
                    po_status = lookup_po_milestone(current_po_id, page, auth_headers)
                    po_status_cache[current_po_id] = po_status
                    time.sleep(0.3)
                except Exception as e:
                    print(f"error: {e}")
                    po_status_cache[current_po_id] = "Error"

            po_status = po_status_cache[current_po_id]
            if col_status:
                ws.cell(row=row_idx, column=col_status).value = po_status
                stats["status_updated"] += 1
            print(f"{label}  {po_status}")

        elif not current_ir_id and current_po_id and col_status:
            po_status = lookup_po_status_from_bulk(current_po_id, pogrid_index)
            if po_status:
                ws.cell(row=row_idx, column=col_status).value = po_status
                stats["status_updated"] += 1
                print(f"{label}  {po_status}  (bulk status)")

    return stats

# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------

def main():
    registry_path = get_registry_path()
    status_mode = get_status_mode()

    if not registry_path.lower().endswith((".xlsx", ".xlsm")):
        print(f"Error: selected file is not an Excel workbook:\n  {registry_path}")
        return

    if not os.path.isfile(registry_path):
        print(f"Error: registry file not found:\n  {registry_path}")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"Loading : {os.path.basename(registry_path)}")
    wb = load_workbook(registry_path)

    os.makedirs(SESSION_DIR, exist_ok=True)

    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                SESSION_DIR,
                channel="chrome",
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
                ignore_https_errors=True,
            )
        except Exception:
            context = p.chromium.launch_persistent_context(
                SESSION_DIR,
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
                ignore_https_errors=True,
            )

        page = context.new_page()
        print("Checking session...")
        auth_headers = ensure_authenticated(page, context)

        try:
            page.evaluate("() => { window.moveTo(-10000, 0); }")
        except Exception:
            pass

        # Shared caches across all sheets — avoids duplicate API calls
        ir_cache        = {}   # invoice_str  → (ir_id, po_id)
        cheque_cache    = {}   # ir_id        → (status_str, chk_num, chk_date)
        po_status_cache = {}   # po_id        → status_str

        print(f"\nStatus mode: {status_mode}")
        irgrid_index = preload_irgrid(page, auth_headers)
        pogrid_index = preload_pogrid(page, auth_headers) if status_mode == "fast" else {}

        total = {"filled_po": 0, "filled_ir": 0, "filled_cheque": 0,
                 "status_updated": 0, "not_found": 0, "bulk_hits": 0,
                 "fallback_lookups": 0, "cheque_issued_skipped": 0}

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            print(f"\n{'='*60}")
            print(f"  Sheet: {sheet_name}")
            print(f"{'='*60}")

            stats = sync_sheet(ws, sheet_name, page, auth_headers,
                               status_mode, irgrid_index, pogrid_index,
                               ir_cache, cheque_cache, po_status_cache)

            for k in total:
                total[k] += stats.get(k, 0)

            print(f"\n  Sheet summary — "
                  f"Status updated: {stats['status_updated']}  "
                  f"PO filled: {stats['filled_po']}  "
                  f"IR filled: {stats['filled_ir']}  "
                  f"Cheque filled: {stats['filled_cheque']}  "
                  f"Cheque issued skipped: {stats['cheque_issued_skipped']}  "
                  f"Not in P2P: {stats['not_found']}  "
                  f"Bulk hits: {stats['bulk_hits']}  "
                  f"Fallback searches: {stats['fallback_lookups']}")

        context.close()

    backup_path = backup_existing_workbook(registry_path)
    wb.save(registry_path)

    print(f"\n{'='*60}")
    print("  TOTAL SUMMARY")
    print(f"{'='*60}")
    print(f"  Status updated  : {total['status_updated']}")
    print(f"  PO filled       : {total['filled_po']}")
    print(f"  IR filled       : {total['filled_ir']}")
    print(f"  Cheque filled   : {total['filled_cheque']}")
    print(f"  Cheque issued skipped: {total['cheque_issued_skipped']}")
    print(f"  Not in P2P yet  : {total['not_found']}")
    print(f"  Bulk IRGrid hits: {total['bulk_hits']}")
    print(f"  Fallback searches: {total['fallback_lookups']}")
    print(f"\n  Backup saved    : {backup_path}")
    print(f"  Workbook updated: {registry_path}")
    print(f"{'='*60}")


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        import traceback
        print(f"\nUnexpected error: {e}")
        traceback.print_exc()
    input("\nPress Enter to close...")
