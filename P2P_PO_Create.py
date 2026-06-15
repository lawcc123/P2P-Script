"""
P2P_PO_Create.py

Creates Yardi P2P purchase orders from a cheque registry workbook.

Expected columns:
  P2P Status
  Contractor Name
  P.O. #
  P2P IR #
  Invoice No
  Invoice Date (DD/MM/YYYY)
  Invoice Amount
  Account Code
  Remarks

The property is inferred from the worksheet name. A Property column is optional
and will be used only when present.

Only rows whose P2P Status contains "require" + "create" + "po" are processed.
The script backs up the original workbook, then saves updates to the same file.

Run:
  python P2P_PO_Create.py --dry-run
  python P2P_PO_Create.py
"""

import base64
import json
import os
import re
import shutil
import sys
import time
from copy import deepcopy
from datetime import datetime
from decimal import Decimal, InvalidOperation

from openpyxl import load_workbook
from playwright.sync_api import sync_playwright


def get_base_dir():
    if getattr(sys, "frozen", False):
        return os.path.dirname(sys.executable)
    return os.path.dirname(os.path.abspath(__file__))


BASE_DIR = get_base_dir()
CONFIG_PATH = os.path.join(BASE_DIR, "p2p_private_config.json")
SESSION_DIR = os.path.join(BASE_DIR, "browser_session")
DEFAULT_INPUT_PATH = os.path.join(BASE_DIR, "01. Cheque Registry_Year 2026.xlsx")
OUTPUT_DIR = os.path.join(BASE_DIR, "Output Excel File")


def load_private_config():
    if not os.path.isfile(CONFIG_PATH):
        raise FileNotFoundError(
            "Missing private config file: p2p_private_config.json. "
            "Copy p2p_private_config.example.json, fill in your private values, "
            "and keep p2p_private_config.json out of public repos."
        )
    with open(CONFIG_PATH, "r", encoding="utf-8") as config_file:
        return json.load(config_file)


PRIVATE_CONFIG = load_private_config()

if not os.path.isfile(DEFAULT_INPUT_PATH):
    search_dirs = [BASE_DIR, os.path.dirname(BASE_DIR)]
    excel_files = []
    for folder in search_dirs:
        if not folder or not os.path.isdir(folder):
            continue
        for filename in os.listdir(folder):
            if filename.startswith("~$"):
                continue
            if filename.lower().endswith((".xlsx", ".xlsm")):
                excel_files.append(os.path.join(folder, filename))
    registry_files = [
        path for path in excel_files
        if "cheque registry" in os.path.basename(path).lower()
    ]
    if registry_files:
        DEFAULT_INPUT_PATH = registry_files[0]
    elif excel_files:
        DEFAULT_INPUT_PATH = excel_files[0]

LOGIN_URL = PRIVATE_CONFIG["login_url"]
API_BASE = PRIVATE_CONFIG["api_base"].rstrip("/")
P2P_DATABASE = PRIVATE_CONFIG["database"]
P2P_ORIGIN = PRIVATE_CONFIG["origin"]
P2P_REFERER = PRIVATE_CONFIG["referer"]
PO_URL = f"{API_BASE}/po"
AUTOCOMPLETE_URL = f"{API_BASE}/scriptingengine/report/filterautocomplete"
PERSON_INFO_URL = f"{API_BASE}/person/getPersonInfo"

EXPENSE_OPERATIONAL = 3680
EXPENSE_RESERVE = 3682
EXPENSE_SHARED_2_WAY = 3685

H_PROPERTY = "Property"
H_STATUS = "P2P Status"
H_VENDOR = "Contractor Name"
H_PO = "P.O. #"
H_INVOICE = "Invoice No"
H_INVOICE_DATE = "Invoice Date (DD/MM/YYYY)"
H_AMOUNT = "Invoice Amount"
H_ACCOUNT = "Account Code"
H_REMARKS = "Remarks"

WARNING_LABELS = {
    "overbudget": "overbudget",
    "over budget": "overbudget",
    "budget": "overbudget",
    "insurance expired": "insurance expired",
    "insurance expire": "insurance expired",
    "insurance": "insurance expired",
}

WARNING_OVERRIDE_FIELDS = {
    "Override": True,
    "OverrideWarning": True,
    "OverrideWarnings": True,
    "IgnoreWarning": True,
    "IgnoreWarnings": True,
    "ForceSave": True,
    "BudgetOverride": True,
    "OverrideBudget": True,
    "OverBudgetOverride": True,
    "InsuranceOverride": True,
    "OverrideInsurance": True,
    "OverrideInsuranceExpired": True,
}


def clean_dropped_path(raw_path):
    path = str(raw_path or "").strip()
    if path.startswith("& "):
        path = path[2:].strip()
    return path.strip().strip('"').strip("'")


def get_input_path():
    path_args = []
    for arg in sys.argv[1:]:
        lowered = arg.lower()
        if lowered == "--dry-run":
            continue
        path_args.append(arg)
    if path_args:
        return clean_dropped_path(" ".join(path_args))

    print("Drag and drop the cheque registry Excel file into this terminal, then press Enter.")
    print(f"Press Enter without a file to use: {os.path.basename(DEFAULT_INPUT_PATH)}")
    selected = clean_dropped_path(input("Excel file: "))
    return selected or DEFAULT_INPUT_PATH


def get_backup_path(input_path):
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    _, ext = os.path.splitext(os.path.basename(input_path))
    return os.path.join(OUTPUT_DIR, f"back up_{ts}{ext}")


def backup_existing_workbook(input_path):
    backup_path = get_backup_path(input_path)
    shutil.copy2(input_path, backup_path)
    return backup_path


def is_dry_run():
    return any(arg.lower() == "--dry-run" for arg in sys.argv[1:])


def get_auth_headers(context):
    all_cookies = {c["name"]: c["value"] for c in context.cookies()}
    jwt_token = all_cookies.get(".JWTAUTH")
    if not jwt_token:
        return None
    try:
        payload_b64 = jwt_token.split(".")[1]
        payload_b64 += "=" * (4 - len(payload_b64) % 4)
        payload = json.loads(base64.urlsafe_b64decode(payload_b64))
        if payload.get("exp", 0) < time.time():
            return None
        login_guid = payload.get("UserLoginGUID", "")
        role = payload.get(
            "http://schemas.microsoft.com/ws/2008/06/identity/claims/role", ""
        )
    except Exception:
        return None
    return {
        "Authorization": f"Bearer {jwt_token}",
        "Database": P2P_DATABASE,
        "Loginguid": login_guid,
        "Role": role,
        "Content-Type": "application/json;charset=UTF-8",
        "Accept": "application/json, text/plain, */*",
        "Accept-Language": "en-US,en;q=0.9",
        "Origin": P2P_ORIGIN,
        "Referer": P2P_REFERER,
        "Sec-Fetch-Dest": "empty",
        "Sec-Fetch-Mode": "cors",
        "Sec-Fetch-Site": "same-origin",
    }


def wait_for_jwt(context, timeout_seconds=300):
    deadline = time.time() + timeout_seconds
    while time.time() < deadline:
        headers = get_auth_headers(context)
        if headers:
            return headers
        time.sleep(1)
    return None


def ensure_authenticated(page, context):
    headers = get_auth_headers(context)
    if headers:
        print("Valid session found - skipping login.\n")
        return headers
    print("\nBrowser opened - please log in to continue.")
    print("After logging in, click 'Elevate (Production)' to open P2P.")
    print("The script will continue automatically once you are in.\n")
    page.goto(LOGIN_URL)
    headers = wait_for_jwt(context, timeout_seconds=300)
    if headers:
        print("Login detected - continuing.\n")
        return headers
    raise TimeoutError("Login timed out after 5 minutes.")


def is_empty(value):
    return value is None or str(value).strip() == ""


def normalize_header(value):
    return str(value or "").strip()


def get_col_map(ws):
    return {
        normalize_header(cell.value): cell.column
        for cell in ws[1]
        if cell.value is not None
    }


def normalize_text(value):
    text = str(value or "").lower()
    text = re.sub(r"[^a-z0-9]+", " ", text)
    return " ".join(text.split())


def cell_text(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def money(value):
    if value is None:
        raise ValueError("missing invoice amount")
    try:
        return float(Decimal(str(value).replace(",", "").strip()))
    except (InvalidOperation, ValueError):
        raise ValueError(f"invalid invoice amount: {value}")


def compact_account_code(value):
    text = cell_text(value)
    if text.endswith(".0"):
        text = text[:-2]
    return text.strip()


def should_process_status(value):
    status = normalize_text(value)
    return "require" in status and "create" in status and "po" in status


def is_contract_status(value):
    return "contract" in normalize_text(value)


def created_status(warnings=None):
    status = f"PO is created on {datetime.now().strftime('%d/%m/%Y')}"
    warning_text = format_warning_status(warnings)
    if warning_text:
        status = f"{status}, {warning_text}"
    return status


def dry_run_status():
    return f"Dry run OK on {datetime.now().strftime('%d/%m/%Y')} - PO not created"


def failed_status(message):
    message = str(message).replace("\r", " ").replace("\n", " ").strip()
    return f"Failed on {datetime.now().strftime('%d/%m/%Y')}: {message[:180]}"


def extract_warning_text(raw_value):
    if raw_value is None:
        return ""
    if isinstance(raw_value, str):
        return raw_value
    if isinstance(raw_value, dict):
        parts = []
        for value in raw_value.values():
            text = extract_warning_text(value)
            if text:
                parts.append(text)
        return " ".join(parts)
    if isinstance(raw_value, list):
        return " ".join(extract_warning_text(item) for item in raw_value)
    return str(raw_value)


def warning_labels_from_text(text):
    normalized = normalize_text(text)
    labels = []
    seen = set()
    for keyword, label in WARNING_LABELS.items():
        if normalize_text(keyword) in normalized and label not in seen:
            labels.append(label)
            seen.add(label)
    return labels


def format_warning_status(warnings):
    if not warnings:
        return ""
    if isinstance(warnings, str):
        labels = warning_labels_from_text(warnings)
    else:
        labels = list(warnings)
    return "/".join(labels)


def autocomplete(page, headers, kind, params):
    resp = page.request.get(
        f"{AUTOCOMPLETE_URL}/{kind}",
        params=params,
        headers=headers,
    )
    if not resp.ok:
        raise RuntimeError(f"{kind} lookup failed: HTTP {resp.status}")
    data = resp.json()
    if not isinstance(data, list):
        raise RuntimeError(f"{kind} lookup returned unexpected data")
    return data


def int_id(value):
    if value is None:
        return 0
    return int(float(value))


def parse_vendor_label(label):
    text = str(label or "")
    match = re.search(r"\((.*?)\)", text)
    inside = match.group(1) if match else text
    return inside.split(",")[0].strip()


def vendor_search_terms(vendor_name):
    text = cell_text(vendor_name)
    terms = []
    seen = set()
    for length in range(len(text), 2, -1):
        term = text[:length].strip()
        key = term.lower()
        if key and key not in seen:
            terms.append(term)
            seen.add(key)
    return terms


def pick_first_vendor_result(results, vendor_name, search_term):
    if not results:
        raise RuntimeError(f"vendor not found: {vendor_name}")
    result = results[0]
    return result, parse_vendor_label(result.get("label", "")), search_term


def parse_person_info(raw_text, fallback_name):
    text = str(raw_text or "")
    first_line = text.splitlines()[0] if text.splitlines() else ""
    payee_code = first_line.split()[0].strip() if first_line.split() else ""
    name_match = re.search(r"/([^/]+)/", first_line)
    vendor_name = name_match.group(1).strip() if name_match else fallback_name
    type_match = re.search(r"/([^/]+)/(\d+)", first_line)
    payee_type = int(type_match.group(2)) if type_match else 5
    return payee_code, vendor_name, payee_type


def lookup_vendor(page, headers, vendor_name, cache):
    key = normalize_text(vendor_name)
    if key in cache:
        return cache[key]

    result = None
    display_name = ""
    matched_search = ""
    for search_term in vendor_search_terms(vendor_name):
        results = autocomplete(
            page,
            headers,
            "vendor_active",
            {"searchstring": search_term},
        )
        if results:
            result, display_name, matched_search = pick_first_vendor_result(
                results, vendor_name, search_term
            )
            break
    if not result:
        raise RuntimeError(f"vendor not found: {vendor_name}")

    vendor_id = int_id(result.get("id"))

    info_resp = page.request.get(f"{PERSON_INFO_URL}/{vendor_id}", headers=headers)
    if not info_resp.ok:
        raise RuntimeError(f"vendor info lookup failed: HTTP {info_resp.status}")
    vendor_info = info_resp.text()
    payee_code, clean_name, payee_type = parse_person_info(vendor_info, display_name)

    vendor = {
        "id": vendor_id,
        "name": clean_name,
        "info": vendor_info,
        "payee_code": payee_code,
        "payee_type": payee_type,
        "matched_search": matched_search,
    }
    cache[key] = vendor
    return vendor


def property_search_text(value):
    text = cell_text(value)
    norm_text = normalize_text(text)
    if "bgsf" in norm_text or "bgate" in norm_text or "bamburgh" in norm_text or "shared" in norm_text:
        return "bgate"
    digits = re.findall(r"\d+", text)
    if digits:
        return digits[-1]
    return text


def parse_property_desc(label, fallback):
    text = str(label or "")
    match = re.search(r"\((.*?)\)", text)
    inside = match.group(1) if match else text
    desc = inside.split(",")[0].strip()
    return desc or cell_text(fallback)


def pick_property_result(results, property_value):
    search = property_search_text(property_value)
    norm_search = normalize_text(search)
    candidates = []
    for item in results:
        label = str(item.get("label", ""))
        label4 = str(item.get("label4", ""))
        norm_all = normalize_text(f"{label} {label4}")
        if norm_search and norm_search in norm_all:
            candidates.append(item)
    if not candidates:
        raise RuntimeError(f"property not found: {property_value}")
    return candidates[0]


def lookup_property(page, headers, property_value, cache):
    key = normalize_text(property_value)
    if key in cache:
        return cache[key]

    search = property_search_text(property_value)
    results = autocomplete(page, headers, "property", {"searchstring": search})
    result = pick_property_result(results, property_value)
    prop = {
        "id": int_id(result.get("id")),
        "desc": parse_property_desc(result.get("label"), property_value),
    }
    cache[key] = prop
    return prop


def property_value_for_row(ws, col_map, row_idx):
    if H_PROPERTY in col_map:
        value = row_value(ws, row_idx, col_map, H_PROPERTY)
        if not is_empty(value):
            return value

    sheet = normalize_text(ws.title)
    if "1136" in sheet:
        return "1136"
    if "1254" in sheet:
        return "1254"
    if "bgsf" in sheet or "bgate" in sheet or "bamburgh" in sheet or "shared" in sheet:
        return "bgate"
    raise ValueError(f"cannot infer property from worksheet name: {ws.title}")


def expense_type_for(property_value, account_code):
    prop = normalize_text(property_value)
    account = compact_account_code(account_code)
    account_base = account.split("-")[0].strip()

    if "bgsf" in prop or "bgate" in prop or "bamburgh" in prop or "shared" in prop:
        return EXPENSE_SHARED_2_WAY
    if ("1136" in prop or "1254" in prop) and account_base == "2810":
        return EXPENSE_RESERVE
    return EXPENSE_OPERATIONAL


def expense_type_name(expense_type_id):
    names = {
        EXPENSE_OPERATIONAL: "Operational Expense",
        EXPENSE_RESERVE: "Reserve Expense",
        EXPENSE_SHARED_2_WAY: "Shared Facilities 2 Way",
    }
    return names.get(expense_type_id, "")


def parse_account_desc(label):
    match = re.search(r"\((.*?)\)", str(label or ""))
    return match.group(1).strip() if match else ""


def pick_account_result(results, account_code):
    code = compact_account_code(account_code)
    code_base = code.split("-")[0].strip()
    for item in results:
        label2 = str(item.get("label2", "")).strip()
        if label2 == code or label2.split("-")[0] == code_base:
            return item
    raise RuntimeError(f"account not found: {account_code}")


def lookup_account(page, headers, expense_type_id, property_id, account_code, cache):
    key = (expense_type_id, property_id, compact_account_code(account_code))
    if key in cache:
        return cache[key]

    results = autocomplete(
        page,
        headers,
        "account",
        {
            "expensetype": expense_type_id,
            "property": property_id,
            "searchstring": compact_account_code(account_code),
        },
    )
    result = pick_account_result(results, account_code)
    acct_code = str(result.get("label2", "")).strip()
    acct_desc = parse_account_desc(result.get("label"))
    account = {
        "id": int_id(result.get("id")),
        "code": acct_code,
        "desc": acct_desc,
        "code_and_desc": f"{acct_code} [{acct_desc}]" if acct_desc else acct_code,
    }
    cache[key] = account
    return account


def get_po_template(page, headers):
    try:
        resp = page.request.get(f"{PO_URL}/0", headers=headers)
        if resp.ok:
            data = resp.json()
            if isinstance(data, dict):
                return data
    except Exception:
        pass
    return {}


def zero_date():
    return "0001-01-01T00:00:00"


def now_iso():
    return datetime.now().strftime("%Y-%m-%dT%H:%M:%S")


def build_header_description(remarks, invoice_no):
    description = cell_text(remarks)
    invoice = cell_text(invoice_no)
    if invoice and invoice.lower() not in description.lower():
        if description:
            return f"{description} Inv no. {invoice}"
        return f"Inv no. {invoice}"
    return description or invoice


def build_detail_description(remarks):
    return cell_text(remarks)


def blank_detail():
    return {
        "POId": 0,
        "Id": 0,
        "IsDeleted": False,
        "IsClosed": False,
        "PropertyId": 0,
        "PayAccountId": 0,
        "AccountId": 0,
        "QuantityOrdered": "0",
        "QuantityReceived": 0,
        "UnitPrice": 0,
        "TotalCost": 0,
        "TranUnitPrice": 0,
        "TranTotalCost": 0,
        "TranGross": 0,
        "TaxAmount1": 0,
        "TaxAmount2": 0,
        "TranTaxAmount1": 0,
        "TranTaxAmount2": 0,
        "TaxOpted": True,
        "Receive": False,
    }


def build_detail(template_detail, prop, account, row_description, amount):
    detail = deepcopy(template_detail) if isinstance(template_detail, dict) else blank_detail()
    detail.update(
        {
            "POId": 0,
            "Id": 0,
            "IsDeleted": False,
            "IsClosed": False,
            "PropertyId": prop["id"],
            "PropertyDesc": prop["desc"],
            "AccountId": account["id"],
            "AcctCode": account["code"],
            "AcctDesc": account["desc"],
            "AcctCodeAndDesc": account["code_and_desc"],
            "Description": row_description,
            "QuantityOrdered": "1",
            "QuantityReceived": 0,
            "UnitPrice": amount,
            "TotalCost": amount,
            "TranUnitPrice": amount,
            "TranTotalCost": amount,
            "TranGross": amount,
            "TaxAmount1": 0,
            "TaxAmount2": 0,
            "TranTaxAmount1": 0,
            "TranTaxAmount2": 0,
            "TranTotalSalesTax": 0,
            "TaxOpted": True,
            "Receive": False,
            "ReceivedDate": zero_date(),
            "DateReceived": zero_date(),
            "VendorId": 0,
        }
    )
    return detail


def build_po_payload(
    template,
    vendor,
    prop,
    account,
    expense_type_id,
    amount,
    header_description,
    detail_description,
):
    payload = deepcopy(template) if isinstance(template, dict) else {}
    order_date = now_iso()
    details = payload.get("DetailList") if isinstance(payload.get("DetailList"), list) else []
    template_detail = details[0] if details else {}
    blank_tail = [deepcopy(item) for item in details[1:] if isinstance(item, dict)]
    real_detail = build_detail(template_detail, prop, account, detail_description, amount)

    payload.update(
        {
            "Id": 0,
            "Code": "",
            "Total": amount,
            "TranTotal": amount,
            "TaxTotal": 0,
            "CalcTotal": amount,
            "CalcBase": amount,
            "DetailTotal": amount,
            "DetailTotalAmount": amount,
            "DetailTotalGross": amount,
            "DetailTotalTax1": 0,
            "DetailTotalTax2": 0,
            "VendorId": vendor["id"],
            "Vendor": vendor["name"],
            "VendorName": f"{vendor['name']} ({vendor['payee_code']})",
            "VendorCode": f"{vendor['payee_code']}-{vendor['name']}",
            "VendorInfo": vendor["info"],
            "PayeeInfo": vendor["info"],
            "PayeeType": vendor["payee_type"],
            "payeeCode": vendor["payee_code"],
            "Description": header_description,
            "ExpenseTypeId": expense_type_id,
            "ExpenseType": expense_type_name(expense_type_id),
            "DisplayTypeId": 1,
            "WorkflowId": 0,
            "FromDate": zero_date(),
            "ToDate": zero_date(),
            "PaymentDueDate": zero_date(),
            "ActualDeliveryDate": zero_date(),
            "ScheduledDeliveryDate": zero_date(),
            "CloseDate": zero_date(),
            "BudgetMonth": zero_date(),
            "OrderDate": order_date,
            "RequestedDate": order_date,
            "RequiredByDate": order_date,
            "IsClosed": False,
            "DetailCount": 1,
            "DetailList": [real_detail] + blank_tail,
            "ExpenseTypeList": [
                {"id": 3680, "name": "Operational Expense"},
                {"id": 3682, "name": "Reserve Expense"},
                {"id": 3683, "name": "Cheque Requisition"},
                {"id": 3685, "name": "Shared Facilities 2 Way"},
                {"id": 3686, "name": "Shared Facilities 3 Way"},
                {"id": 3687, "name": "Shared Facilities 4 Way"},
                {"id": 3688, "name": "Preliminary"},
                {"id": 3689, "name": "Service Contracts"},
            ],
            "DisplayTypeList": [{"id": 1, "name": "Standard PO Display Type"}],
            "CurrencyList": [{"id": 1, "name": "CAD"}, {"id": 2, "name": "USD"}],
        }
    )
    return payload


def post_po(page, headers, payload):
    return page.request.post(PO_URL, data=json.dumps(payload), headers=headers)


def po_id_from_response(body):
    try:
        data = json.loads(body)
    except json.JSONDecodeError:
        raise RuntimeError(f"PO create returned non-JSON response: {body[:120]}")
    po_id = data.get("POID") or data.get("PoId") or data.get("POId") or data.get("Id")
    if not po_id:
        raise RuntimeError(f"PO create response missing POID: {data}")
    return int_id(po_id)


def warning_override_payload(payload):
    override_payload = deepcopy(payload)
    override_payload.update(WARNING_OVERRIDE_FIELDS)
    return override_payload


def create_po(page, headers, payload):
    resp = post_po(page, headers, payload)
    body = resp.text()
    if resp.ok:
        return po_id_from_response(body), []

    warning_text = extract_warning_text(body)
    warning_labels = warning_labels_from_text(warning_text)
    if not warning_labels:
        raise RuntimeError(f"PO create failed: HTTP {resp.status} {body[:300]}")

    override_resp = post_po(page, headers, warning_override_payload(payload))
    override_body = override_resp.text()
    if not override_resp.ok:
        raise RuntimeError(
            "PO create failed after warning override: "
            f"HTTP {override_resp.status} {override_body[:300]}"
        )
    return po_id_from_response(override_body), warning_labels


def fetch_po_code(page, headers, po_id):
    resp = page.request.get(f"{PO_URL}/{po_id}", headers=headers)
    if not resp.ok:
        raise RuntimeError(f"created PO readback failed: HTTP {resp.status}")
    data = resp.json()
    code = cell_text(data.get("Code") or data.get("PONumber") or data.get("PO") or po_id)
    return code.strip() or str(po_id)


def row_value(ws, row_idx, col_map, header):
    col = col_map.get(header)
    return ws.cell(row=row_idx, column=col).value if col else None


def set_row_value(ws, row_idx, col_map, header, value):
    col = col_map.get(header)
    if col:
        ws.cell(row=row_idx, column=col).value = value


def validate_headers(col_map):
    required = [
        H_STATUS,
        H_VENDOR,
        H_PO,
        H_INVOICE,
        H_AMOUNT,
        H_ACCOUNT,
        H_REMARKS,
    ]
    missing = [header for header in required if header not in col_map]
    if missing:
        raise RuntimeError(f"missing required columns: {', '.join(missing)}")


def collect_candidate_rows(wb):
    candidates = []
    for ws in wb.worksheets:
        col_map = get_col_map(ws)
        validate_headers(col_map)
        for row_idx in range(2, ws.max_row + 1):
            status = row_value(ws, row_idx, col_map, H_STATUS)
            po_value = row_value(ws, row_idx, col_map, H_PO)
            if not should_process_status(status):
                continue
            if is_contract_status(status):
                continue
            if not is_empty(po_value):
                continue
            candidates.append((ws, col_map, row_idx))
    return candidates


def process_row(ws, col_map, row_idx, page, headers, caches, template, dry_run=False):
    property_value = property_value_for_row(ws, col_map, row_idx)
    vendor_name = row_value(ws, row_idx, col_map, H_VENDOR)
    invoice_no = row_value(ws, row_idx, col_map, H_INVOICE)
    amount_value = row_value(ws, row_idx, col_map, H_AMOUNT)
    account_code = row_value(ws, row_idx, col_map, H_ACCOUNT)
    remarks = row_value(ws, row_idx, col_map, H_REMARKS)

    if is_empty(vendor_name):
        raise ValueError("missing contractor name")
    if is_empty(account_code):
        raise ValueError("missing account code")

    amount = money(amount_value)
    header_description = build_header_description(remarks, invoice_no)
    detail_description = build_detail_description(remarks)
    if is_empty(detail_description):
        raise ValueError("missing remarks/description")

    prop = lookup_property(page, headers, property_value, caches["property"])
    expense_type_id = expense_type_for(property_value, account_code)
    vendor = lookup_vendor(page, headers, cell_text(vendor_name), caches["vendor"])
    account = lookup_account(
        page,
        headers,
        expense_type_id,
        prop["id"],
        account_code,
        caches["account"],
    )
    payload = build_po_payload(
        template,
        vendor,
        prop,
        account,
        expense_type_id,
        amount,
        header_description,
        detail_description,
    )
    if dry_run:
        set_row_value(ws, row_idx, col_map, H_STATUS, dry_run_status())
        return {
            "property": f"{prop['desc']} ({prop['id']})",
            "vendor": f"{vendor['name']} ({vendor['id']})",
            "vendor_search": vendor.get("matched_search", ""),
            "expense_type": f"{expense_type_name(expense_type_id)} ({expense_type_id})",
            "account": f"{account['code']} ({account['id']})",
            "amount": amount,
            "header_description": header_description,
            "detail_description": detail_description,
            "detail_count": len(payload.get("DetailList", [])),
        }

    po_id, warnings = create_po(page, headers, payload)
    po_code = fetch_po_code(page, headers, po_id)

    set_row_value(ws, row_idx, col_map, H_PO, po_code)
    set_row_value(ws, row_idx, col_map, H_STATUS, created_status(warnings))
    return po_code


def wait_for_enter_before_exit():
    try:
        input("\nPress Enter to close this window...")
    except (EOFError, KeyboardInterrupt):
        pass


def main():
    dry_run = is_dry_run()
    input_path = get_input_path()
    if not input_path.lower().endswith((".xlsx", ".xlsm")):
        print(f"Error: selected file is not an Excel workbook:\n  {input_path}")
        return
    if not os.path.isfile(input_path):
        print(f"Error: PO create file not found:\n  {input_path}")
        return

    os.makedirs(OUTPUT_DIR, exist_ok=True)

    print(f"Loading: {os.path.basename(input_path)}")
    wb = load_workbook(input_path)
    candidates = collect_candidate_rows(wb)
    if not candidates:
        print("No rows found with P2P Status requiring PO creation.")
        return

    if dry_run:
        print(f"\nRows ready for dry run: {len(candidates)}")
        print("Dry run will perform live lookups and build payloads, but will not create POs.")
        confirm = input("Type DRYRUN to continue: ").strip()
        if confirm != "DRYRUN":
            print("Cancelled - no POs created.")
            return
    else:
        print(f"\nRows ready to create PO: {len(candidates)}")
        confirm = input("Type CREATE to create these POs in Yardi: ").strip()
        if confirm != "CREATE":
            print("Cancelled - no POs created.")
            return

    os.makedirs(SESSION_DIR, exist_ok=True)
    caches = {"vendor": {}, "property": {}, "account": {}}

    with sync_playwright() as p:
        try:
            context = p.chromium.launch_persistent_context(
                SESSION_DIR,
                channel="chrome",
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
                ignore_https_errors=True,
            )
        except Exception:
            context = p.chromium.launch_persistent_context(
                SESSION_DIR,
                headless=False,
                args=["--disable-blink-features=AutomationControlled"],
                ignore_https_errors=True,
            )

        page = context.new_page()
        print("Checking session...")
        headers = ensure_authenticated(page, context)
        template = get_po_template(page, headers)

        created = 0
        dry_run_ok = 0
        failed = 0
        for ws, col_map, row_idx in candidates:
            invoice = cell_text(row_value(ws, row_idx, col_map, H_INVOICE))
            label = f"[{ws.title}] Row {row_idx} | Invoice {invoice}"
            print(f"{label} ...", end=" ", flush=True)
            try:
                result = process_row(
                    ws, col_map, row_idx, page, headers, caches, template, dry_run
                )
                if dry_run:
                    dry_run_ok += 1
                    print(
                        "dry run OK | "
                        f"{result['property']} | {result['vendor']} | "
                        f"Search: {result['vendor_search']} | "
                        f"{result['expense_type']} | {result['account']} | "
                        f"${result['amount']:.2f} | "
                        f"Header: {result['header_description']} | "
                        f"Detail: {result['detail_description']}"
                    )
                else:
                    created += 1
                    print(f"created PO {result}")
                time.sleep(0.4)
            except Exception as exc:
                failed += 1
                set_row_value(ws, row_idx, col_map, H_STATUS, failed_status(exc))
                print(f"failed: {exc}")

        context.close()

    backup_path = backup_existing_workbook(input_path)
    wb.save(input_path)
    print("\nDone.")
    if dry_run:
        print(f"  Dry run OK: {dry_run_ok}")
    else:
        print(f"  Created   : {created}")
    print(f"  Failed    : {failed}")
    print(f"  Backup    : {backup_path}")
    print(f"  Updated   : {input_path}")


if __name__ == "__main__":
    try:
        main()
    finally:
        wait_for_enter_before_exit()
